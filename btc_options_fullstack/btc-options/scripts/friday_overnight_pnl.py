"""
Friday-overnight short-ATM-straddle backtest for Jan–Mar 2026.

For every Friday in the window:
  • Open at 23:30 IST a short straddle on the next-day (Saturday) expiry,
    strike = ATM at 23:30 (rounded to 100/500 grid available in chain).
  • Mark to market every minute through 00:30 IST Saturday.
  • Close at 00:30 IST Saturday.
  • Apply slippage (slippage.ts model) on entry and exit only.
  • Apply brokerage (brokerage.ts offer rate + 18% GST) on entry and exit.
  • Apply margin engine (mirrors marginEngine.ts) using IV from
    backend/app/core/greeks.py::implied_vol — same path as Historical chain.

Output: scripts/friday_overnight_pnl.xlsx
  • Sheet "Summary"  — one row per (Friday × lot size)
  • Sheet "<Date>"   — per-minute MTM curve (one sheet per Friday)
"""
from __future__ import annotations
import os, sys, glob
from datetime import datetime, timezone, timedelta

import duckdb

sys.path.insert(0, os.path.abspath("backend"))
sys.path.insert(0, os.path.abspath("."))
from app.core.greeks import implied_vol, compute_greeks
from scripts.margin_engine import (
    CONTRACT_VALUE, MarginLeg, compute_portfolio_margin, implied_forward_pcp,
)

CV  = CONTRACT_VALUE
IST = timezone(timedelta(hours=5, minutes=30))
RFR = 0.0

# ── Cost models (verbatim ports of frontend/src/utils/{slippage,brokerage}.ts) ──
def slip_per_side(mark, hour_ist, is_weekend, qty):
    if mark <= 0: return 0
    hour_mult    = 1.10 if hour_ist < 6 else (1.05 if hour_ist < 9 else 1.00)
    weekend_mult = 1.05 if is_weekend else 1.0
    qty_btc      = max(0.001, qty * CV)
    return max(0.10/qty_btc + 3.85, 0.012*mark) * hour_mult * weekend_mult * CV

def brokerage(spot, mark, qty):
    return min(0.0001*spot, 0.035*mark) * qty * CV * 1.18

# ── Historical-chain identical T (settle 12:00 UTC = 17:30 IST, 365-day year) ──
def historical_T(ts_unix: int, expiry_yyyy_mm_dd: str) -> float:
    expiry_dt  = datetime.strptime(expiry_yyyy_mm_dd, "%Y-%m-%d").replace(
        tzinfo=timezone.utc, hour=12)
    cur_dt = datetime.fromtimestamp(ts_unix, tz=timezone.utc)
    return max(0.0001, (expiry_dt - cur_dt).total_seconds() / (365*24*3600))

LOT_SIZES = [100, 500, 1000, 2000]

# ── Date setup ──────────────────────────────────────────────────────────────
fridays = []
for month in (1, 2, 3):
    for day in range(1, 32):
        try:
            d = datetime(2026, month, day, tzinfo=IST)
            if d.weekday() == 4:  # Friday
                fridays.append(d.date())
        except ValueError:
            break

con = duckdb.connect()
SPOT_PARQ = "/home/abhis/btc-data/data/spot/BTCUSD_1min.parquet"
OPT_BASE  = "/home/abhis/btc-data/data/options"

def has_expiry(date_yyyy_mm_dd: str) -> bool:
    return os.path.isdir(f"{OPT_BASE}/expiry={date_yyyy_mm_dd}")

def get_chain_strikes(expiry_yyyy_mm_dd: str) -> list[int]:
    p = f"{OPT_BASE}/expiry={expiry_yyyy_mm_dd}"
    return sorted(int(d.split("=")[1]) for d in os.listdir(p) if d.startswith("strike="))

def fetch_minute_marks(expiry_str, strike, kind, ts_from, ts_to):
    path = f"{OPT_BASE}/expiry={expiry_str}/strike={strike}/{kind}.parquet"
    if not os.path.exists(path): return None
    return con.execute(
        f"""SELECT timestamp_unix, mark_close FROM read_parquet('{path}')
            WHERE timestamp_unix >= {ts_from} AND timestamp_unix <= {ts_to}
            ORDER BY timestamp_unix"""
    ).df()

def fetch_minute_spot(ts_from, ts_to):
    return con.execute(
        f"""SELECT timestamp_unix, mark_close FROM read_parquet('{SPOT_PARQ}')
            WHERE timestamp_unix >= {ts_from} AND timestamp_unix <= {ts_to}
            ORDER BY timestamp_unix"""
    ).df()

# ── Run analysis ──────────────────────────────────────────────────────────────
summary_rows = []
detail_sheets = {}   # date_str → list of dicts (per-minute rows)

for fd in fridays:
    sat = fd + timedelta(days=1)
    expiry_str = sat.strftime("%Y-%m-%d")
    if not has_expiry(expiry_str):
        print(f"  [skip] {fd}: no Saturday expiry parquet ({expiry_str})")
        continue

    entry_dt = datetime(fd.year, fd.month, fd.day, 23, 30, tzinfo=IST)
    exit_dt  = entry_dt + timedelta(hours=1)
    entry_ts = int(entry_dt.timestamp())
    exit_ts  = int(exit_dt.timestamp())

    # spot
    spot_df = fetch_minute_spot(entry_ts, exit_ts)
    if spot_df.empty:
        print(f"  [skip] {fd}: no spot data in window"); continue
    spot_e  = float(spot_df.iloc[0]["mark_close"])
    spot_x  = float(spot_df.iloc[-1]["mark_close"])

    # ATM strike — closest available in the chain
    strikes = get_chain_strikes(expiry_str)
    if not strikes:
        print(f"  [skip] {fd}: no strikes for {expiry_str}"); continue
    K = min(strikes, key=lambda s: abs(s - spot_e))

    # Marks at entry/exit + per-minute MTM
    ce_df = fetch_minute_marks(expiry_str, K, "CE", entry_ts, exit_ts)
    pe_df = fetch_minute_marks(expiry_str, K, "PE", entry_ts, exit_ts)
    if ce_df is None or pe_df is None or ce_df.empty or pe_df.empty:
        print(f"  [skip] {fd}: missing CE/PE for K={K}"); continue

    mark_e_ce = float(ce_df.iloc[0]["mark_close"])
    mark_e_pe = float(pe_df.iloc[0]["mark_close"])
    mark_x_ce = float(ce_df.iloc[-1]["mark_close"])
    mark_x_pe = float(pe_df.iloc[-1]["mark_close"])

    # IV + Greeks at entry (historical chain identical math)
    T_e = historical_T(entry_ts, expiry_str)
    iv_e_ce = implied_vol(mark_e_ce, spot_e, K, T_e, RFR, "call")
    iv_e_pe = implied_vol(mark_e_pe, spot_e, K, T_e, RFR, "put")
    g_e_ce = compute_greeks(spot_e, K, T_e, RFR, iv_e_ce if iv_e_ce>0 else 0.5, "call")
    g_e_pe = compute_greeks(spot_e, K, T_e, RFR, iv_e_pe if iv_e_pe>0 else 0.5, "put")

    F = implied_forward_pcp(mark_e_ce, mark_e_pe, K)

    # Build per-minute MTM curve
    spot_map = {int(r.timestamp_unix): float(r.mark_close) for r in spot_df.itertuples()}
    ce_map   = {int(r.timestamp_unix): float(r.mark_close) for r in ce_df.itertuples()}
    pe_map   = {int(r.timestamp_unix): float(r.mark_close) for r in pe_df.itertuples()}

    minute_rows = []
    common_ts = sorted(set(spot_map) & set(ce_map) & set(pe_map))

    # ── Per-lot summary + detail ────────────────────────────────────────────
    for qty in LOT_SIZES:
        # Costs (entry-side only locked at the trade time)
        slip_in = (slip_per_side(mark_e_ce, 23, False, qty)
                 + slip_per_side(mark_e_pe, 23, False, qty)) * qty
        slip_out_default = (slip_per_side(mark_x_ce, 0, True, qty)
                          + slip_per_side(mark_x_pe, 0, True, qty)) * qty
        brok_in  = brokerage(spot_e, mark_e_ce, qty) + brokerage(spot_e, mark_e_pe, qty)
        brok_out_default = brokerage(spot_x, mark_x_ce, qty) + brokerage(spot_x, mark_x_pe, qty)

        # Margin (engine, IV from historical chain path)
        legs = [
            MarginLeg(strike=K, is_call=True,  is_buy=False, qty=qty,
                      current_price=mark_e_ce*CV, iv=iv_e_ce, T=T_e, forward=F),
            MarginLeg(strike=K, is_call=False, is_buy=False, qty=qty,
                      current_price=mark_e_pe*CV, iv=iv_e_pe, T=T_e, forward=F),
        ]
        mr = compute_portfolio_margin(legs, spot_e)
        margin = mr.portfolio_margin if mr else 0.0

        entry_credit = (mark_e_ce + mark_e_pe) * CV * qty

        # Per-minute MTM (assumes no exit costs for floating MTM — only at close)
        max_mtm = -1e18; max_mtm_ts = entry_ts; min_mtm = +1e18; min_mtm_ts = entry_ts
        for ts in common_ts:
            ce, pe = ce_map[ts], pe_map[ts]
            mtm_premium = (ce + pe) * CV * qty   # cost to buy back
            unreal_pnl  = entry_credit - mtm_premium
            net_mtm     = unreal_pnl - slip_in - brok_in   # cost to enter only
            if net_mtm > max_mtm: max_mtm = net_mtm; max_mtm_ts = ts
            if net_mtm < min_mtm: min_mtm = net_mtm; min_mtm_ts = ts
            if qty == LOT_SIZES[0]:
                minute_rows.append({"ts_ist": datetime.fromtimestamp(ts, IST).strftime("%H:%M:%S"),
                                    "spot": round(spot_map[ts], 2),
                                    "ce_mark": round(ce, 2),
                                    "pe_mark": round(pe, 2)})

        # Closed exit P&L
        exit_debit = (mark_x_ce + mark_x_pe) * CV * qty
        gross      = entry_credit - exit_debit
        net_close  = gross - slip_in - slip_out_default - brok_in - brok_out_default

        max_mtm_dt = datetime.fromtimestamp(max_mtm_ts, IST).strftime("%H:%M:%S")
        min_mtm_dt = datetime.fromtimestamp(min_mtm_ts, IST).strftime("%H:%M:%S")

        summary_rows.append({
            "Friday": str(fd),
            "Expiry (Sat)": expiry_str,
            "Qty (lots)": qty,
            "Strike (ATM)": K,
            "Spot @ Entry": round(spot_e, 2),
            "Spot @ Exit":  round(spot_x, 2),
            "Spot Δ":       round(spot_x - spot_e, 2),
            "Entry Time IST": "23:30:00",
            "Exit Time IST":  "00:30:00",
            "Entry CE Mark $/BTC": round(mark_e_ce, 2),
            "Entry PE Mark $/BTC": round(mark_e_pe, 2),
            "Entry CE δ":  round(g_e_ce.delta, 4),
            "Entry PE δ":  round(g_e_pe.delta, 4),
            "Entry CE IV%": round(iv_e_ce*100, 2),
            "Entry PE IV%": round(iv_e_pe*100, 2),
            "Exit CE Mark $/BTC": round(mark_x_ce, 2),
            "Exit PE Mark $/BTC": round(mark_x_pe, 2),
            "Margin $": round(margin, 2),
            "Entry Credit $":   round(entry_credit, 2),
            "Exit Debit $":     round(exit_debit, 2),
            "Gross P&L $":      round(gross, 2),
            "Slip Entry $":     round(slip_in, 2),
            "Slip Exit $":      round(slip_out_default, 2),
            "Brok Entry $":     round(brok_in, 2),
            "Brok Exit $":      round(brok_out_default, 2),
            "Net P&L (Close) $": round(net_close, 2),
            "Net P&L % of Margin": round(net_close/margin*100, 2) if margin>0 else 0,
            "Max MTM Time IST": max_mtm_dt,
            "Max MTM $":        round(max_mtm, 2),
            "Min MTM Time IST": min_mtm_dt,
            "Min MTM $":        round(min_mtm, 2),
        })

    detail_sheets[str(fd)] = {
        "minute_rows": minute_rows,
        "K": K, "expiry": expiry_str,
        "entry_credit_per_lot": (mark_e_ce + mark_e_pe) * CV,
        "slip_in_per_lot":  (slip_per_side(mark_e_ce, 23, False, LOT_SIZES[0])
                          +  slip_per_side(mark_e_pe, 23, False, LOT_SIZES[0])),
        "brok_in_per_lot":  (brokerage(spot_e, mark_e_ce, 1) + brokerage(spot_e, mark_e_pe, 1)),
    }
    print(f"  ✓ {fd}  K={K}  CE_iv={iv_e_ce*100:.1f}% PE_iv={iv_e_pe*100:.1f}%  "
          f"close_net@1000lots=${[r['Net P&L (Close) $'] for r in summary_rows if r['Friday']==str(fd) and r['Qty (lots)']==1000][0]:+.2f}")

# ── Write Excel ────────────────────────────────────────────────────────────
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Summary"
if not summary_rows:
    ws["A1"] = "No data"
else:
    headers = list(summary_rows[0].keys())
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")
    for r, row in enumerate(summary_rows, 2):
        for c, h in enumerate(headers, 1):
            ws.cell(r, c, row[h])
    # Auto-fit (rough)
    for c, h in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = max(len(str(h))+2, 14)
    ws.freeze_panes = "A2"

# Per-Friday detail sheets
for date_str, info in detail_sheets.items():
    ws = wb.create_sheet(date_str)
    ws["A1"] = f"Friday {date_str} → Sat {info['expiry']}  K={info['K']}"
    ws["A1"].font = Font(bold=True, size=12)
    cols = ["Time IST", "Spot $", "CE Mark $/BTC", "PE Mark $/BTC",
            "Combined $/BTC", "MTM/lot $", "Net MTM/lot $ (after entry costs)"]
    for c, h in enumerate(cols, 1):
        cell = ws.cell(3, c, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="305496")
    entry_combined = info["entry_credit_per_lot"]   # per lot (already × CV)
    entry_costs_per_lot = info["slip_in_per_lot"] + info["brok_in_per_lot"]   # per lot
    for r, row in enumerate(info["minute_rows"], 4):
        combined = row["ce_mark"] + row["pe_mark"]
        mtm_per_lot = entry_combined - combined * CV
        net_mtm_per_lot = mtm_per_lot - entry_costs_per_lot
        ws.cell(r, 1, row["ts_ist"])
        ws.cell(r, 2, row["spot"])
        ws.cell(r, 3, row["ce_mark"])
        ws.cell(r, 4, row["pe_mark"])
        ws.cell(r, 5, round(combined, 2))
        ws.cell(r, 6, round(mtm_per_lot, 4))
        ws.cell(r, 7, round(net_mtm_per_lot, 4))
    for c in range(1, len(cols)+1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 16
    ws.freeze_panes = "A4"

out_path = os.path.abspath("scripts/friday_overnight_pnl.xlsx")
wb.save(out_path)
print(f"\n  📊 Saved: {out_path}")
print(f"     Sheets: Summary + {len(detail_sheets)} per-Friday detail sheets")
